#!/usr/bin/env python3
"""
run_backtest_on_sids_data.py
Processes the actual EDGAR data Sid collected, pulls Yahoo Finance
prices, reconstructs outcomes, and runs the full SCP backtest.
"""
import sys, os
sys.path.insert(0, '/home/claude/autocall-trap')

import numpy as np
import openpyxl
from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import Optional, List
import csv

# Since we can't reach Yahoo Finance from sandbox, we'll work with
# what we have and set up the pipeline to run locally

INPUT_FILE = '/mnt/user-data/uploads/autocall_trap_backtest_data.xlsx'

def load_sids_data():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    notes = []
    for row in range(2, ws.max_row + 1):
        issue_date = ws.cell(row=row, column=4).value
        mat_date = ws.cell(row=row, column=6).value
        
        if isinstance(issue_date, datetime):
            issue_str = issue_date.strftime('%Y-%m-%d')
        else:
            issue_str = str(issue_date)
        
        if isinstance(mat_date, datetime) and isinstance(issue_date, datetime):
            mat_yrs = round((mat_date - issue_date).days / 365.25, 2)
        else:
            mat_yrs = 2.0
        
        freq = ws.cell(row=row, column=9).value
        memory_val = ws.cell(row=row, column=13).value
        memory = str(memory_val).strip().lower() in ('yes', 'true', '1')
        
        # Determine first autocall obs (default: 1 for most, 2 if quarterly with delay)
        first_ac = 1
        
        note = {
            'note_id': f"{ws.cell(row=row, column=2).value[:2].upper()}-{ws.cell(row=row, column=3).value}-{issue_str[:4]}{issue_str[5:7]}",
            'issuer': ws.cell(row=row, column=2).value,
            'underlying': ws.cell(row=row, column=3).value,
            'issue_date': issue_str,
            'S0': float(ws.cell(row=row, column=5).value),
            'par': 1000.0,
            'maturity': mat_yrs,
            'n_obs': int(ws.cell(row=row, column=7).value),
            'coupon_rate': float(ws.cell(row=row, column=8).value),
            'autocall_trigger': float(ws.cell(row=row, column=10).value),
            'coupon_barrier': float(ws.cell(row=row, column=11).value),
            'ki_barrier': float(ws.cell(row=row, column=12).value),
            'memory': memory,
            'first_autocall_obs': first_ac,
            'risk_free_rate': 0.045,  # Will be overwritten by FRED
            'atm_iv': 0.30,  # Will be overwritten by realized vol estimate
            'div_yield': 0.01,  # Default, adjusted per ticker
            'issuer_estimated_value': float(ws.cell(row=row, column=14).value) if ws.cell(row=row, column=14).value else 0,
            'maturity_date': mat_date.strftime('%Y-%m-%d') if isinstance(mat_date, datetime) else '',
            'freq': freq,
        }
        notes.append(note)
    
    return notes


def export_for_backtest(notes, filepath):
    """Export to CSV format for the backtest engine."""
    fieldnames = [
        'note_id', 'issuer', 'underlying', 'issue_date', 'S0', 'par',
        'maturity', 'n_obs', 'coupon_rate', 'autocall_trigger',
        'coupon_barrier', 'ki_barrier', 'memory', 'first_autocall_obs',
        'risk_free_rate', 'atm_iv', 'div_yield', 'issuer_estimated_value',
    ]
    
    with open(filepath, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for n in notes:
            row = {k: n.get(k, '') for k in fieldnames}
            writer.writerow(row)
    
    print(f"Exported {len(notes)} notes to {filepath}")


def print_summary(notes):
    print(f"\n{'DATASET SUMMARY':=^65}")
    print(f"  Total notes:     {len(notes)}")
    
    # By issuer
    issuers = {}
    for n in notes:
        issuers[n['issuer']] = issuers.get(n['issuer'], 0) + 1
    print(f"\n  By issuer:")
    for k, v in sorted(issuers.items()):
        print(f"    {k}: {v}")
    
    # By underlying type
    tickers = {}
    for n in notes:
        tickers[n['underlying']] = tickers.get(n['underlying'], 0) + 1
    print(f"\n  By underlying ({len(tickers)} unique):")
    for k, v in sorted(tickers.items()):
        print(f"    {k}: {v}")
    
    # Maturity distribution
    mats = [n['maturity'] for n in notes]
    print(f"\n  Maturity range: {min(mats):.1f}y - {max(mats):.1f}y")
    
    # Estimated value / margin
    evs = [n['issuer_estimated_value'] for n in notes if n['issuer_estimated_value'] > 0]
    margins = [(1000 - ev) / 10 for ev in evs]
    print(f"\n  Estimated value range: ${min(evs):.0f} - ${max(evs):.0f}")
    print(f"  Issuer-admitted margin: {min(margins):.1f}% - {max(margins):.1f}% (avg {np.mean(margins):.1f}%)")
    
    # Matured vs live
    now = datetime(2026, 3, 17)
    matured = sum(1 for n in notes 
                  if n['maturity_date'] and datetime.strptime(n['maturity_date'], '%Y-%m-%d') < now)
    print(f"\n  Matured: {matured}")
    print(f"  Still live: {len(notes) - matured}")


if __name__ == '__main__':
    print("=" * 65)
    print("PROCESSING SID'S EDGAR DATA")
    print("=" * 65)
    
    notes = load_sids_data()
    print_summary(notes)
    
    output_path = '/home/claude/autocall-trap/data/sids_notes.csv'
    os.makedirs('/home/claude/autocall-trap/data', exist_ok=True)
    export_for_backtest(notes, output_path)
    
    print(f"\n{'NEXT STEPS':=^65}")
    print(f"  On your machine, run:")
    print(f"  python data_pipeline.py --input data/sids_notes.csv")
    print(f"  python -m src.backtest --csv notes_universe.csv --paths 100000")
    print(f"{'='*65}")
